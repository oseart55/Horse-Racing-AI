import os
import requests
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import traceback

FILE_NAME = "predictions.xlsx"
SHEET_NAME = "Predictions"
COLUMNS = ["Date", "Track", "Race Number", "Prediction", "Probability", "Correct"]
HEADERS = {
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_2 like Mac OS X) "
                  "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.2 "
                  "Mobile/15E148 Safari/604.1"
}
# --- Helper functions ---
def autosize_columns(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(file_path)


def log_exception(exc, context="", logfile="errors.log"):
    timestamp = date.today().strftime("%Y-%m-%d %H:%M:%S")
    entry = (
        f"\n[{timestamp}] EXCEPTION\n"
        f"Context: {context}\n"
        f"Type: {type(exc).__name__}\n"
        f"Message: {exc}\n"
        f"Traceback:\n{traceback.format_exc()}\n"
        f"{'-'*60}\n"
    )
    os.makedirs(os.path.dirname(logfile), exist_ok=True) if os.path.dirname(logfile) else None
    with open(logfile, "a", encoding="utf-8") as f:
        f.write(entry)


def first_horse(race_details, pool_key):
    pools = race_details.get(pool_key)
    if isinstance(pools, list) and len(pools) > 0:
        return pools[0].get("horseName")
    return None


# --- Append a new prediction to Excel ---
def append_prediction(track, race_number, prediction, probability, race_date=None, file_path=FILE_NAME):
    if race_date is None:
        race_date = date.today()

    df_row = pd.DataFrame(
        [[race_date, track, race_number, prediction, probability, None]],
        columns=COLUMNS
    )

    if not os.path.exists(file_path):
        df_row.to_excel(file_path, sheet_name=SHEET_NAME, index=False)
        autosize_columns(file_path, SHEET_NAME)
        return

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        if SHEET_NAME in writer.book.sheetnames:
            sheet = writer.book[SHEET_NAME]
            startrow = sheet.max_row
            df_row.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=startrow)
        else:
            df_row.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    autosize_columns(file_path, SHEET_NAME)

def getResults(brisCode: str, raceNumber: int):
    today_str = date.today().strftime("%Y-%m-%d")
    resultsUrl = f"https://www.twinspires.com/api/raceresults/results/{today_str}/{brisCode}/Thoroughbred/{raceNumber}"
    results = requests.get(resultsUrl, headers=HEADERS)
    if results.status_code == 200:
        return results.json()
    
# --- Update all rows with missing Correct ---
def update_correct_column(file_path=FILE_NAME, sheet_name=SHEET_NAME):
    if not os.path.exists(file_path):
        return  # nothing to do

    df = pd.read_excel(file_path, sheet_name=sheet_name)

    for idx, row in df.iterrows():
        if str(row.get("Correct")).upper() in ["Y", "N"]:
            continue

        track = row.get("Track")
        race_number = row.get("Race Number")
        prediction = row.get("Prediction")

        try:
            raceResults = getResults(track, race_number)
            race_details = raceResults.get("raceDetails") or {}

            winPool = first_horse(race_details, "winPools")
            placePool = first_horse(race_details, "placePools")
            showPool = first_horse(race_details, "showPools")

            df.at[idx, "Correct"] = "Y" if prediction in [winPool, placePool, showPool] else "F"

        except Exception as e:
            log_exception(e, f"Error processing {track} race {race_number}")
            df.at[idx, "Correct"] = None

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    autosize_columns(file_path, sheet_name)


# --- Unified function ---
def append_and_update(track, race_number, prediction, probability, race_date=None, file_path=FILE_NAME):
    update_correct_column(file_path)
    append_prediction(track, race_number, prediction, probability, race_date, file_path)
    
